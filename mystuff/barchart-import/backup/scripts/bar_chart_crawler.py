import csv
import glob
import os
from datetime import datetime

from openpyxl import load_workbook
from scripts.constants import *
from scripts.crawler import *
from scripts.utils import *


def parse_file_from_bar_chart():
    files = glob.glob(DOWNLOAD_FOLDER + "*.csv")

    wb = load_workbook(filename=OUTPUT_FOLDER + DATA_BAR_CHART_FILE_NAME)
    sheet = wb.active
    # date_format_xls = out_wb.add_format({'num_format': 'MM/DD/y'})  # Full format: YYYY-MM-DD HH:DD:SS

    with open(files[0]) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=",")
        line_count = get_first_empty_row(sheet)
        for row in csv_reader:
            if line_count == 1:
                for column_i, column_data in enumerate(row):
                    sheet.cell(row=line_count, column=column_i + 1, value=row[column_i])
                line_count += 1
            else:
                if (
                    "Downloaded from " in row[0]
                ):  # This is check for file end (every file ends with this)
                    continue
                for column_i, column_data in enumerate(row):
                    if check_is_date(column_data, "%y/%m/%d"):
                        date_obj = datetime.strptime(column_data, "%y/%m/%d")
                        cell = sheet.cell(
                            row=line_count, column=column_i + 1, value=date_obj
                        )
                        cell.number_format = "DD//MM/YYYY"
                    if check_is_digit(column_data):
                        sheet.cell(
                            row=line_count,
                            column=column_i + 1,
                            value=float(column_data),
                        )
                    else:
                        sheet.cell(
                            row=line_count, column=column_i + 1, value=column_data
                        )

                sheet.cell(
                    row=line_count, column=column_i + 2, value=datetime.now().date()
                )
                line_count += 1
        print(f"Processed {line_count} lines from bar chart file")

    wb.save(filename=OUTPUT_FOLDER + DATA_BAR_CHART_FILE_NAME)
    # Delete downloaded files
    for file in files:
        os.remove(file)


def bar_chart_file_fill_from_yahoo():
    print("Opening file " + OUTPUT_FOLDER + DATA_BAR_CHART_FILE_NAME)
    wb = load_workbook(filename=OUTPUT_FOLDER + DATA_BAR_CHART_FILE_NAME)
    sheet = wb.active
    print("Reading " + str(sheet.max_row) + " rows")

    for row in range(2, sheet.max_row + 1):
        for column in "A":  # Here you can add or reduce the columns
            date_crawled = sheet.cell(row=row, column=16).value
            diff = datetime.today() - date_crawled

            if sheet.cell(row=row, column=17).value is None:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value  # the value of the specific cell
                parsed_url = YAHOO_URL.replace(
                    "{0}", cell.replace(".", "")
                )  # Prepare URL to be crawled
                r = requests.get(url=parsed_url)  # Make the actual request
                data = r.json()  # Parse result as json

                if data is not None:
                    try:
                        if r.status_code == 404:
                            sector = "[NOT FOUND]"
                            market_cap = "[NOT FOUND]"
                            trailing_pe = "[NOT FOUND]"
                            eps = "[NOT FOUND]"
                        else:
                            sector = data["quoteSummary"]["result"][0]["assetProfile"][
                                "sector"
                            ]
                            market_cap = data["quoteSummary"]["result"][0][
                                "summaryDetail"
                            ]["marketCap"]["raw"]
                            trailing_pe = data["quoteSummary"]["result"][0][
                                "summaryDetail"
                            ]["trailingPE"]["raw"]
                            eps = data["quoteSummary"]["result"][0][
                                "defaultKeyStatistics"
                            ]["trailingEps"]["raw"]
                    except (KeyError, TimeoutError):
                        sector = "[NONE]"
                        market_cap = "[NONE]"
                        trailing_pe = "[NONE]"
                        eps = "[NONE]"
                else:
                    sector = "[NOT CONNECTED]"
                    market_cap = "[NOT CONNECTED]"
                    trailing_pe = "[NOT CONNECTED]"
                    eps = "[NOT CONNECTED]"
                sheet.cell(row=row, column=17, value=sector)
                sheet.cell(row=row, column=18, value=market_cap)
                sheet.cell(row=row, column=19, value=trailing_pe)
                sheet.cell(row=row, column=20, value=eps)
                print(
                    "Updating data row for " + cell + " crawled on " + str(date_crawled)
                )

            # T-1
            if sheet.cell(row=row, column=21).value is None and diff.days >= 2:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value
                t1 = crawl_old_data(cell.replace(".", ""), date_crawled, 1)
                print("Adding column for T-1 for " + cell + " value:" + str(t1))
                sheet.cell(row=row, column=21, value=t1)

            # T-2
            if sheet.cell(row=row, column=22).value is None and diff.days >= 3:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value
                t2 = crawl_old_data(cell.replace(".", ""), date_crawled, 2)
                print("Adding column for T-2 for " + cell + " value:" + str(t2))
                sheet.cell(row=row, column=22, value=t2)

            # T-5
            if sheet.cell(row=row, column=23).value is None and diff.days >= 6:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value
                t5 = crawl_old_data(cell.replace(".", ""), date_crawled, 5)
                print("Adding column for T-5 for " + cell + " value:" + str(t5))
                sheet.cell(row=row, column=23, value=t5)

            # T-10
            if sheet.cell(row=row, column=24).value is None and diff.days >= 11:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value
                t10 = crawl_old_data(cell.replace(".", ""), date_crawled, 10)
                print("Adding column for T-10 for " + cell + " value:" + str(t10))
                sheet.cell(row=row, column=24, value=t10)
    wb.save(filename=OUTPUT_FOLDER + DATA_BAR_CHART_FILE_NAME)
    print("Saving file " + OUTPUT_FOLDER + DATA_BAR_CHART_FILE_NAME)
    pass
