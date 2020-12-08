import requests
from openpyxl import load_workbook
from scripts.crawler import *
from scripts.utils import *


def download_data_from_chameleon():
    crawler = Crawler()
    driver = crawler.driver
    driver.get("https://marketchameleon.com/Account/Login")

    user_input_button = driver.find_element_by_css_selector("#UserName")
    user_input_button.send_keys(BAR_CHART_USERNAME)

    pass_input_button = driver.find_element_by_css_selector("#Password")
    pass_input_button.send_keys(BAR_CHART_PASSWORD)

    time.sleep(2)
    login_button = driver.find_element_by_css_selector(
        "#site-body-inner > div.site-login-outer > div:nth-child(1) > form > div:nth-child(5) > input"
    )
    login_button.click()
    time.sleep(5)
    driver.set_window_size(1920, 1080)
    driver.get(MARKET_CHAMELEON_URL)
    time.sleep(5)
    driver.find_element_by_css_selector(MARKET_CHAMELEON_NEXT_BUTTON_ID)
    data = [[]]
    exit_crawling_page = False
    while True:
        time.sleep(3)
        table = driver.find_element_by_css_selector(
            "#opt_unusual_volume > tbody:nth-child(2)"
        )
        rows = table.find_elements_by_xpath("*")
        for row in rows:
            tds = row.find_elements_by_tag_name("td")
            data_row = [
                tds[0].text,
                tds[1].find_element_by_class_name("mplink").text,
                tds[2].text,
                tds[3].text,
                tds[4].text,
                tds[5].text,
                tds[6].text,
                tds[7].text,
                tds[8].text,
                tds[9].text,
                tds[10].text,
                tds[11].text,
                tds[12].text,
                tds[13].text,
                tds[14].text,
                tds[15].text,
                tds[16].text,
                tds[17].text,
            ]
            data.append(data_row)
            print(data_row[2])

        while not get_next_button(driver):
            pass
        find_next_button(MARKET_CHAMELEON_NEXT_BUTTON_ID, driver)

        if exit_crawling_page:
            break

        exit_crawling_page = check_is_last_page(driver)
        # next_page = None
    time.sleep(3)
    driver.close()
    print("Chameleon website crawled successfully")
    return data


def parse_data_into_file(data):
    wb = load_workbook(filename=OUTPUT_FOLDER + DATA_CHAMELEON_FILE_NAME)
    sheet = wb.active
    line_count = get_first_empty_row(sheet)

    for row in data:
        column = 0
        for column_i, column_data in enumerate(row):
            if check_is_date(column_data, "%y/%m/%d"):
                date_obj = datetime.strptime(column_data, "%y/%m/%d")
                cell = sheet.cell(row=line_count, column=column_i + 1, value=date_obj)
                cell.number_format = "DD//MM/YYYY"
            if check_is_digit(column_data):
                sheet.cell(
                    row=line_count,
                    column=column_i + 1,
                    value=float(column_data.replace(",", "")),
                )
            else:
                sheet.cell(row=line_count, column=column_i + 1, value=column_data)
            column = column_i
        if column != 0:
            sheet.cell(row=line_count, column=column + 2, value=datetime.now().date())
        line_count += 1
    wb.save(filename=OUTPUT_FOLDER + DATA_CHAMELEON_FILE_NAME)
    print(f"Processed {line_count} lines from chameleon website")


def chameleon_file_fill_from_yahoo():
    wb = load_workbook(filename=OUTPUT_FOLDER + DATA_CHAMELEON_FILE_NAME)
    sheet = wb.active
    print(str(sheet.max_row) + " rows")

    for row in range(3, sheet.max_row + 1):
        for column in "B":  # Here you can add or reduce the columns
            date_crawled = sheet.cell(row=row, column=19).value
            diff = datetime.today() - date_crawled
            if sheet.cell(row=row, column=20).value is None:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value  # the value of the specific cell
                print(cell)
                parsed_url = YAHOO_URL.replace(
                    "{0}", cell.replace(".", "")
                )  # Prepare URL to be crawled
                r = requests.get(url=parsed_url)  # Make the actual request
                data = r.json()  # Parse result as json

                if data is not None:
                    try:
                        if r.status_code == 404:
                            sector = "[NOT FOUND]"
                            market_cap = "[NOT FOUND]"
                            trailing_pe = "[NOT FOUND]"
                            eps = "[NOT FOUND]"
                        else:
                            sector = data["quoteSummary"]["result"][0]["assetProfile"][
                                "sector"
                            ]
                            market_cap = data["quoteSummary"]["result"][0][
                                "summaryDetail"
                            ]["marketCap"]["raw"]
                            trailing_pe = data["quoteSummary"]["result"][0][
                                "summaryDetail"
                            ]["trailingPE"]["raw"]
                            eps = data["quoteSummary"]["result"][0][
                                "defaultKeyStatistics"
                            ]["trailingEps"]["raw"]
                    except (KeyError, TimeoutError):
                        sector = "[NONE]"
                        market_cap = "[NONE]"
                        trailing_pe = "[NONE]"
                        eps = "[NONE]"
                else:
                    sector = "[NOT CONNECTED]"
                    market_cap = "[NOT CONNECTED]"
                    trailing_pe = "[NOT CONNECTED]"
                    eps = "[NOT CONNECTED]"
                sheet.cell(row=row, column=20, value=sector)
                sheet.cell(row=row, column=21, value=market_cap)
                sheet.cell(row=row, column=22, value=trailing_pe)
                sheet.cell(row=row, column=23, value=eps)
                print(
                    "Updating data row for " + cell + " crawled on " + str(date_crawled)
                )

            # T-1
            if sheet.cell(row=row, column=24).value is None and diff.days >= 2:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value
                t1 = crawl_old_data(cell.replace(".", ""), date_crawled, 1)
                print("Adding column for T-1 for " + cell + " value:" + str(t1))
                sheet.cell(row=row, column=24, value=t1)

            # T-2
            if sheet.cell(row=row, column=25).value is None and diff.days >= 3:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value
                t2 = crawl_old_data(cell.replace(".", ""), date_crawled, 2)
                print("Adding column for T-2 for " + cell + " value:" + str(t2))
                sheet.cell(row=row, column=25, value=t2)

            # T-5
            if sheet.cell(row=row, column=26).value is None and diff.days >= 6:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value
                t5 = crawl_old_data(cell.replace(".", ""), date_crawled, 5)
                print("Adding column for T-5 for " + cell + " value:" + str(t5))
                sheet.cell(row=row, column=26, value=t5)

            # T-10
            if sheet.cell(row=row, column=27).value is None and diff.days >= 11:
                cell_name = "{}{}".format(column, row)
                cell = sheet[cell_name].value
                t10 = crawl_old_data(cell.replace(".", ""), date_crawled, 10)
                print("Adding column for T-10 for " + cell + " value:" + str(t10))
                sheet.cell(row=row, column=27, value=t10)
    wb.save(filename=OUTPUT_FOLDER + DATA_CHAMELEON_FILE_NAME)
